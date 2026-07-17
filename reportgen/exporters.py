"""
多格式导出器模块

支持 HTML / PDF / Excel / Markdown / JSON 五种格式的导出。
PDF 使用 weasyprint（可选），缺失时自动降级提示。
"""

import io
import json
import warnings
from abc import ABC, abstractmethod
from datetime import datetime
from typing import Any, Dict, List, Optional

# ──────────────────────────────────────────────
# HTML 导出
# ──────────────────────────────────────────────


class HTMLExporter:
    """HTML 格式导出器"""

    def export(self, content: str, output_path: str) -> str:
        """
        将 HTML 内容导出到文件

        Args:
            content: HTML 字符串
            output_path: 输出文件路径

        Returns:
            输出文件的绝对路径
        """
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        return output_path


# ──────────────────────────────────────────────
# PDF 导出
# ──────────────────────────────────────────────

class PDFExporter:
    """PDF 格式导出器（依赖 weasyprint）"""

    def __init__(self):
        self._available = None
        self._html_exporter = HTMLExporter()

    @property
    def available(self) -> bool:
        """检查 weasyprint 是否可用"""
        if self._available is None:
            try:
                from weasyprint import HTML  # noqa: F401
                self._available = True
            except ImportError:
                self._available = False
        return self._available

    def export(self, content: str, output_path: str) -> str:
        """
        将 HTML 内容导出为 PDF

        先渲染为临时 HTML，再通过 weasyprint 转为 PDF。
        如果 weasyprint 不可用，则保存为 HTML 并给出提示。

        Args:
            content: HTML 字符串
            output_path: 输出文件路径（推荐 .pdf）

        Returns:
            输出文件的绝对路径
        """
        if not self.available:
            # 降级处理：保存为 HTML
            html_path = output_path.replace(".pdf", ".html")
            self._html_exporter.export(content, html_path)
            warnings.warn(
                "PDF导出需要安装weasyprint。已保存为HTML文件。\n"
                "安装方法: pip install weasyprint"
            )
            return html_path

        try:
            from weasyprint import HTML

            HTML(string=content).write_pdf(output_path)
            return output_path
        except Exception as e:
            raise RuntimeError(f"PDF 导出失败: {e}")


# ──────────────────────────────────────────────
# Excel 导出
# ──────────────────────────────────────────────

class ExcelExporter:
    """Excel 格式导出器（依赖 openpyxl）"""

    def __init__(self):
        self._available = None

    @property
    def available(self) -> bool:
        if self._available is None:
            try:
                import openpyxl  # noqa: F401
                self._available = True
            except ImportError:
                self._available = False
        return self._available

    def export(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        columns: Optional[List[str]] = None,
        title: str = "报告",
        chart_image: Optional[str] = None,
    ) -> str:
        """
        将数据导出为 Excel 文件

        Args:
            data: 数据行列表
            output_path: 输出文件路径
            columns: 列名列表，默认从数据中自动提取
            title: 工作表标题
            chart_image: base64 编码的图表图片（可选）

        Returns:
            输出文件的绝对路径

        Raises:
            RuntimeError: openpyxl 不可用时
        """
        if not self.available:
            raise RuntimeError(
                "Excel导出需要安装openpyxl。安装方法: pip install openpyxl"
            )

        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]  # Excel 工作表名最长 31 字符

        # ── 样式定义 ──
        header_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        header_alignment = Alignment(horizontal="center", vertical="center")
        cell_font = Font(name="Microsoft YaHei", size=10)
        cell_alignment = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style="thin", color="D9D9D9"),
            right=Side(style="thin", color="D9D9D9"),
            top=Side(style="thin", color="D9D9D9"),
            bottom=Side(style="thin", color="D9D9D9"),
        )

        # ── 标题行 ──
        if not data:
            ws.cell(row=1, column=1, value="无数据")
            wb.save(output_path)
            return output_path

        if not columns:
            columns = list(data[0].keys())

        # 写表头
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        # 写数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, col_name in enumerate(columns, 1):
                value = row_data.get(col_name)
                if value is None:
                    value = ""
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border

        # 自动调整列宽
        for col_idx, col_name in enumerate(columns, 1):
            max_length = len(str(col_name))
            for row_idx in range(2, len(data) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

        # 冻结首行
        ws.freeze_panes = "A2"

        # 如果有图表图片（base64），写入注释或图片
        if chart_image:
            import base64

            try:
                img_data = base64.b64decode(chart_image)
                img_stream = io.BytesIO(img_data)
                from openpyxl.drawing.image import Image

                img = Image(img_stream)
                img.width = 600
                img.height = 300
                img_anchor = "A" + str(len(data) + 3)
                ws.add_image(img, img_anchor)
            except Exception:
                pass  # 图片插入失败不影响主数据

        wb.save(output_path)
        return output_path


# ──────────────────────────────────────────────
# Markdown 导出
# ──────────────────────────────────────────────

class MarkdownExporter:
    """Markdown 格式导出器"""

    def export(self, content: str, output_path: str) -> str:
        """
        导出为 Markdown 文件

        Args:
            content: Markdown 字符串
            output_path: 输出文件路径

        Returns:
            输出文件的绝对路径
        """
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(content)
        return output_path


# ──────────────────────────────────────────────
# JSON 导出
# ──────────────────────────────────────────────

class JSONExporter:
    """JSON 格式导出器"""

    def export(
        self,
        data: List[Dict[str, Any]],
        output_path: str,
        metadata: Optional[Dict[str, Any]] = None,
        **kwargs,
    ) -> str:
        """
        导出为 JSON 文件

        Args:
            data: 数据
            output_path: 输出路径
            metadata: 元数据（可选）
            **kwargs: 传递给 json.dump 的额外参数

        Returns:
            输出文件的绝对路径
        """
        output = {"data": data}
        if metadata:
            output["metadata"] = metadata
        output["exported_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, ensure_ascii=False, indent=2, **kwargs)
        return output_path


# ──────────────────────────────────────────────
# 导出器工厂
# ──────────────────────────────────────────────

class ExporterFactory:
    """导出器工厂类"""

    _exporters = {
        "html": HTMLExporter,
        "pdf": PDFExporter,
        "excel": ExcelExporter,
        "markdown": MarkdownExporter,
        "json": JSONExporter,
    }

    @classmethod
    def create(cls, fmt: str):
        """
        创建对应格式的导出器

        Args:
            fmt: 格式名称 (html / pdf / excel / markdown / json)

        Returns:
            导出器实例

        Raises:
            ValueError: 不支持的格式
        """
        fmt = fmt.lower()
        if fmt not in cls._exporters:
            supported = list(cls._exporters.keys())
            raise ValueError(
                f"不支持的输出格式 '{fmt}'，支持格式: {supported}"
            )
        return cls._exporters[fmt]()

    @classmethod
    def list_formats(cls) -> List[str]:
        """列出所有支持的格式"""
        return list(cls._exporters.keys())
